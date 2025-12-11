"""
报告生成器
生成Excel、HTML、Markdown等格式的综合报告。
"""
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Union, Any, Tuple
import logging
import json
import yaml

from ..utils.file_utils import FileUtils
from ..config.defaults import get_config

logger = logging.getLogger(__name__)


class ReportGenerator:
    """综合报告生成器"""
    
    def __init__(self, config: Optional[Dict] = None):
        """
        初始化报告生成器
        
        Args:
            config: 配置字典
        """
        self.config = get_config(config)
        self.template_dir = Path(__file__).parent / "templates"
        
        logger.info("报告生成器初始化完成")
    
    def generate_all_reports(
        self,
        indels_df: Optional[pd.DataFrame] = None,
        primers_df: Optional[pd.DataFrame] = None,
        blast_results_df: Optional[pd.DataFrame] = None,
        visualization_files: Optional[Dict] = None,
        output_dir: Union[str, Path] = "./reports",
        config: Optional[Dict] = None
    ) -> Dict[str, str]:
        """
        生成所有格式的报告
        
        Args:
            indels_df: Indel数据框
            primers_df: 引物数据框
            blast_results_df: BLAST结果数据框
            visualization_files: 可视化文件路径字典
            output_dir: 输出目录
            config: 配置字典
            
        Returns:
            Dict[str, str]: 报告文件路径字典
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        if config:
            self.config = get_config(config)
        
        reports = {}
        
        # 生成Excel报告
        try:
            excel_report = self.generate_excel_report(
                indels_df, primers_df, blast_results_df, output_dir
            )
            if excel_report:
                reports['excel'] = excel_report
        except Exception as e:
            logger.error(f"生成Excel报告失败: {e}")
        
        # 生成HTML报告
        try:
            html_report = self.generate_html_report(
                indels_df, primers_df, blast_results_df, 
                visualization_files, output_dir
            )
            if html_report:
                reports['html'] = html_report
        except Exception as e:
            logger.error(f"生成HTML报告失败: {e}")
        
        # 生成Markdown报告
        try:
            md_report = self.generate_markdown_report(
                indels_df, primers_df, blast_results_df,
                visualization_files, output_dir
            )
            if md_report:
                reports['markdown'] = md_report
        except Exception as e:
            logger.error(f"生成Markdown报告失败: {e}")
        
        # 生成JSON报告
        try:
            json_report = self.generate_json_report(
                indels_df, primers_df, blast_results_df,
                visualization_files, output_dir
            )
            if json_report:
                reports['json'] = json_report
        except Exception as e:
            logger.error(f"生成JSON报告失败: {e}")
        
        # 生成摘要报告
        try:
            summary_report = self.generate_summary_report(
                indels_df, primers_df, blast_results_df,
                output_dir
            )
            if summary_report:
                reports['summary'] = summary_report
        except Exception as e:
            logger.error(f"生成摘要报告失败: {e}")
        
        logger.info(f"生成了 {len(reports)} 个报告文件")
        return reports
    
    def generate_excel_report(
        self,
        indels_df: Optional[pd.DataFrame],
        primers_df: Optional[pd.DataFrame],
        blast_results_df: Optional[pd.DataFrame],
        output_dir: Union[str, Path]
    ) -> Optional[str]:
        """
        生成Excel报告
        
        Args:
            indels_df: Indel数据框
            primers_df: 引物数据框
            blast_results_df: BLAST结果数据框
            output_dir: 输出目录
            
        Returns:
            str: Excel报告文件路径
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("未安装openpyxl，跳过Excel报告生成")
            return None
        
        output_dir = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_dir / f"qtl_primer_report_{timestamp}.xlsx"
        
        # 创建Workbook
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认的sheet
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 工作表1: 结果摘要
        ws_summary = wb.create_sheet(title="结果摘要")
        
        summary_data = [
            ["QTL精细定位引物设计报告", "", "", ""],
            ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""],
            ["", "", "", ""],
            ["统计信息", "", "", ""],
            ["项目", "数量", "说明", ""],
        ]
        
        if indels_df is not None:
            summary_data.append(["候选Indel", len(indels_df), "过滤后的Indel变异", ""])
        
        if primers_df is not None:
            successful_primers = len(primers_df[primers_df['QUALITY'] != 'FAILED']) if 'QUALITY' in primers_df.columns else len(primers_df)
            summary_data.append(["设计引物", len(primers_df), f"成功设计: {successful_primers}", ""])
        
        if blast_results_df is not None:
            if 'Specificity_Grade' in blast_results_df.columns:
                grade_a = len(blast_results_df[blast_results_df['Specificity_Grade'] == 'A'])
                summary_data.append(["特异性验证", len(blast_results_df), f"A级引物: {grade_a}", ""])
        
        # 写入摘要数据
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=cell_value)
                
                if row_idx == 1:  # 标题
                    cell.font = Font(bold=True, size=14)
                elif row_idx <= 3 or row_idx == 4:  # 标题行
                    cell.font = Font(bold=True)
                    if row_idx == 4:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                
                cell.border = cell_border
        
        # 调整列宽
        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column].width = adjusted_width
        
        # 工作表2: Indel列表
        if indels_df is not None and not indels_df.empty:
            ws_indels = wb.create_sheet(title="候选Indel")
            self._dataframe_to_sheet(indels_df, ws_indels, header_fill, header_font, header_alignment, cell_border)
        
        # 工作表3: 引物设计
        if primers_df is not None and not primers_df.empty:
            ws_primers = wb.create_sheet(title="引物设计")
            self._dataframe_to_sheet(primers_df, ws_primers, header_fill, header_font, header_alignment, cell_border)
        
        # 工作表4: 特异性验证
        if blast_results_df is not None and not blast_results_df.empty:
            ws_blast = wb.create_sheet(title="特异性验证")
            self._dataframe_to_sheet(blast_results_df, ws_blast, header_fill, header_font, header_alignment, cell_border)
            
            # 添加条件格式（如果存在特异性评分）
            if 'Specificity_Score' in blast_results_df.columns:
                from openpyxl.formatting.rule import ColorScaleRule
                
                # 创建颜色刻度规则
                color_scale_rule = ColorScaleRule(
                    start_type='min', start_color='FF0000',  # 红色
                    mid_type='percentile', mid_value=50, mid_color='FFFF00',  # 黄色
                    end_type='max', end_color='00FF00'  # 绿色
                )
                
                # 找到评分列的列号
                score_col_idx = list(blast_results_df.columns).index('Specificity_Score') + 1
                score_col_letter = get_column_letter(score_col_idx)
                
                # 应用条件格式（从第2行开始，第1行是标题）
                ws_blast.conditional_formatting.add(
                    f"{score_col_letter}2:{score_col_letter}{len(blast_results_df)+1}",
                    color_scale_rule
                )
        
        # 保存Excel文件
        wb.save(str(excel_file))
        logger.info(f"Excel报告已生成: {excel_file}")
        
        return str(excel_file)
    
    def _dataframe_to_sheet(self, df: pd.DataFrame, ws, header_fill, header_font, 
                           header_alignment, cell_border):
        """将DataFrame写入Excel工作表"""
        # 写入标题行
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(column_name))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = cell_border
        
        # 写入数据
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                cell.border = cell_border
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
    
    def generate_html_report(
        self,
        indels_df: Optional[pd.DataFrame],
        primers_df: Optional[pd.DataFrame],
        blast_results_df: Optional[pd.DataFrame],
        visualization_files: Optional[Dict],
        output_dir: Union[str, Path]
    ) -> Optional[str]:
        """
        生成HTML报告
        
        Args:
            indels_df: Indel数据框
            primers_df: 引物数据框
            blast_results_df: BLAST结果数据框
            visualization_files: 可视化文件路径字典
            output_dir: 输出目录
            
        Returns:
            str: HTML报告文件路径
        """
        try:
            from jinja2 import Template
        except ImportError:
            logger.error("未安装Jinja2，跳过HTML报告生成")
            return None
        
        output_dir = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        html_file = output_dir / f"qtl_primer_report_{timestamp}.html"
        
        # 准备模板数据
        template_data = {
            'report_title': 'QTL精细定位引物设计报告',
            'generation_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'tool_version': '1.0.0',
            
            # 统计信息
            'indel_count': len(indels_df) if indels_df is not None else 0,
            'primer_count': len(primers_df) if primers_df is not None else 0,
            'blast_count': len(blast_results_df) if blast_results_df is not None else 0,
            
            # 可视化文件
            'visualization_files': visualization_files or {},
            
            # 数据摘要
            'indels_summary': self._get_dataframe_summary(indels_df, 'indels'),
            'primers_summary': self._get_dataframe_summary(primers_df, 'primers'),
            'blast_summary': self._get_dataframe_summary(blast_results_df, 'blast'),
        }
        
        # 如果有BLAST结果，计算特异性统计
        if blast_results_df is not None and not blast_results_df.empty:
            if 'Specificity_Grade' in blast_results_df.columns:
                grade_counts = blast_results_df['Specificity_Grade'].value_counts().to_dict()
                template_data['specificity_stats'] = {
                    'grade_counts': grade_counts,
                    'total': len(blast_results_df)
                }
            
            if 'Specificity_Score' in blast_results_df.columns:
                avg_score = blast_results_df['Specificity_Score'].mean()
                max_score = blast_results_df['Specificity_Score'].max()
                min_score = blast_results_df['Specificity_Score'].min()
                
                template_data['score_stats'] = {
                    'average': f"{avg_score:.2f}",
                    'max': f"{max_score:.2f}",
                    'min': f"{min_score:.2f}"
                }
        
        # 加载模板
        template_path = self.template_dir / "report_template.html"
        if template_path.exists():
            with open(template_path, 'r', encoding='utf-8') as f:
                template_content = f.read()
        else:
            # 使用默认模板
            template_content = self._get_default_html_template()
        
        # 渲染模板
        template = Template(template_content)
        html_content = template.render(**template_data)
        
        # 保存HTML文件
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML报告已生成: {html_file}")
        return str(html_file)
    
    def _get_dataframe_summary(self, df: Optional[pd.DataFrame], df_type: str) -> Dict:
        """获取DataFrame的摘要信息"""
        if df is None or df.empty:
            return {'has_data': False}
        
        summary = {
            'has_data': True,
            'row_count': len(df),
            'column_count': len(df.columns),
            'columns': list(df.columns),
            'sample_rows': df.head(5).to_dict('records')
        }
        
        # 添加类型特定的统计
        if df_type == 'primers' and 'QUALITY' in df.columns:
            quality_counts = df['QUALITY'].value_counts().to_dict()
            summary['quality_stats'] = quality_counts
        
        elif df_type == 'blast' and 'Specificity_Grade' in df.columns:
            grade_counts = df['Specificity_Grade'].value_counts().to_dict()
            summary['grade_stats'] = grade_counts
        
        return summary
    
    def _get_default_html_template(self) -> str:
        """获取默认的HTML模板"""
        return """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ report_title }}</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f5f5f5;
        }
        
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }
        
        .header {
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 2px solid #333;
            padding-bottom: 20px;
        }
        
        .header h1 {
            color: #333;
            margin-bottom: 10px;
        }
        
        .metadata {
            display: flex;
            justify-content: space-between;
            margin-bottom: 20px;
            padding: 15px;
            background-color: #f8f9fa;
            border-radius: 5px;
        }
        
        .section {
            margin-bottom: 30px;
            padding: 20px;
            border: 1px solid #ddd;
            border-radius: 5px;
        }
        
        .section-title {
            color: #2c3e50;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 1px solid #eee;
        }
        
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
        }
        
        .stat-card {
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            text-align: center;
        }
        
        .stat-value {
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
        }
        
        .stat-label {
            font-size: 14px;
            color: #6c757d;
        }
        
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        
        th {
            background-color: #2c3e50;
            color: white;
            padding: 12px;
            text-align: left;
        }
        
        td {
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }
        
        tr:hover {
            background-color: #f5f5f5;
        }
        
        .visualization-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
        }
        
        .viz-item {
            text-align: center;
        }
        
        .viz-item img {
            max-width: 100%;
            height: auto;
            border: 1px solid #ddd;
            border-radius: 5px;
        }
        
        .footer {
            text-align: center;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            color: #6c757d;
            font-size: 14px;
        }
        
        @media (max-width: 768px) {
            .container {
                padding: 15px;
            }
            
            .metadata {
                flex-direction: column;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{{ report_title }}</h1>
            <p>生成时间: {{ generation_time }} | 工具版本: {{ tool_version }}</p>
        </div>
        
        <div class="metadata">
            <div>候选Indel数量: {{ indel_count }}</div>
            <div>设计引物数量: {{ primer_count }}</div>
            <div>BLAST验证数量: {{ blast_count }}</div>
        </div>
        
        {% if specificity_stats %}
        <div class="section">
            <h2 class="section-title">特异性验证统计</h2>
            <div class="stats-grid">
                {% for grade, count in specificity_stats.grade_counts.items() %}
                <div class="stat-card">
                    <div class="stat-value">{{ count }}</div>
                    <div class="stat-label">{{ grade }}级引物</div>
                </div>
                {% endfor %}
            </div>
        </div>
        {% endif %}
        
        {% if indels_summary.has_data %}
        <div class="section">
            <h2 class="section-title">候选Indel (前5行)</h2>
            <table>
                <thead>
                    <tr>
                        {% for column in indels_summary.columns %}
                        <th>{{ column }}</th>
                        {% endfor %}
                    </tr>
                </thead>
                <tbody>
                    {% for row in indels_summary.sample_rows %}
                    <tr>
                        {% for column in indels_summary.columns %}
                        <td>{{ row[column] }}</td>
                        {% endfor %}
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% endif %}
        
        {% if primers_summary.has_data %}
        <div class="section">
            <h2 class="section-title">引物设计结果 (前5行)</h2>
            <table>
                <thead>
                    <tr>
                        {% for column in primers_summary.columns %}
                        <th>{{ column }}</th>
                        {% endfor %}
                    </tr>
                </thead>
                <tbody>
                    {% for row in primers_summary.sample_rows %}
                    <tr>
                        {% for column in primers_summary.columns %}
                        <td>{{ row[column] }}</td>
                        {% endfor %}
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% endif %}
        
        {% if visualization_files %}
        <div class="section">
            <h2 class="section-title">可视化结果</h2>
            <div class="visualization-grid">
                {% for viz_name, viz_path in visualization_files.items() %}
                <div class="viz-item">
                    <img src="{{ viz_path }}" alt="{{ viz_name }}">
                    <p>{{ viz_name }}</p>
                </div>
                {% endfor %}
            </div>
        </div>
        {% endif %}
        
        <div class="footer">
            <p>QTL Primer Toolkit 生成 | 仅供科研参考</p>
        </div>
    </div>
</body>
</html>
        """
    
    def generate_markdown_report(
        self,
        indels_df: Optional[pd.DataFrame],
        primers_df: Optional[pd.DataFrame],
        blast_results_df: Optional[pd.DataFrame],
        visualization_files: Optional[Dict],
        output_dir: Union[str, Path]
    ) -> Optional[str]:
        """
        生成Markdown报告
        
        Args:
            indels_df: Indel数据框
            primers_df: 引物数据框
            blast_results_df: BLAST结果数据框
            visualization_files: 可视化文件路径字典
            output_dir: 输出目录
            
        Returns:
            str: Markdown报告文件路径
        """
        output_dir = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        md_file = output_dir / f"qtl_primer_report_{timestamp}.md"
        
        # 构建Markdown内容
        lines = []
        
        lines.append("# QTL精细定位引物设计报告")
        lines.append("")
        lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"**工具版本**: 1.0.0")
        lines.append("")
        
        lines.append("## 1. 结果摘要")
        lines.append("")
        lines.append("| 项目 | 数量 | 说明 |")
        lines.append("|------|------|------|")
        
        if indels_df is not None:
            lines.append(f"| 候选Indel | {len(indels_df)} | 过滤后的Indel变异 |")
        
        if primers_df is not None:
            successful_primers = len(primers_df[primers_df['QUALITY'] != 'FAILED']) if 'QUALITY' in primers_df.columns else len(primers_df)
            lines.append(f"| 设计引物 | {len(primers_df)} | 成功设计: {successful_primers} |")
        
        if blast_results_df is not None:
            if 'Specificity_Grade' in blast_results_df.columns:
                grade_a = len(blast_results_df[blast_results_df['Specificity_Grade'] == 'A'])
                lines.append(f"| 特异性验证 | {len(blast_results_df)} | A级引物: {grade_a} |")
            else:
                lines.append(f"| 特异性验证 | {len(blast_results_df)} | - |")
        
        lines.append("")
        
        # 特异性统计
        if blast_results_df is not None and not blast_results_df.empty:
            if 'Specificity_Grade' in blast_results_df.columns:
                lines.append("## 2. 特异性验证统计")
                lines.append("")
                lines.append("| 等级 | 数量 | 百分比 |")
                lines.append("|------|------|--------|")
                
                grade_counts = blast_results_df['Specificity_Grade'].value_counts()
                total = len(blast_results_df)
                
                for grade, count in grade_counts.items():
                    percentage = (count / total) * 100
                    lines.append(f"| {grade} | {count} | {percentage:.1f}% |")
                
                lines.append("")
        
        # 引物设计结果
        if primers_df is not None and not primers_df.empty:
            lines.append("## 3. 引物设计结果")
            lines.append("")
            
            # 质量统计
            if 'QUALITY' in primers_df.columns:
                quality_counts = primers_df['QUALITY'].value_counts()
                lines.append("**质量分布**:")
                lines.append("")
                
                for quality, count in quality_counts.items():
                    lines.append(f"- {quality}: {count}")
                
                lines.append("")
            
            # 前5个引物
            lines.append("**前5个引物**:")
            lines.append("")
            lines.append("| ID | 左引物 | 右引物 | 产物大小 | 质量 |")
            lines.append("|----|--------|--------|----------|------|")
            
            for idx, row in primers_df.head().iterrows():
                left_primer = row.get('LEFT_PRIMER', '')[:20] + "..." if len(row.get('LEFT_PRIMER', '')) > 20 else row.get('LEFT_PRIMER', '')
                right_primer = row.get('RIGHT_PRIMER', '')[:20] + "..." if len(row.get('RIGHT_PRIMER', '')) > 20 else row.get('RIGHT_PRIMER', '')
                product_size = row.get('PRODUCT_SIZE', '')
                quality = row.get('QUALITY', '')
                
                lines.append(f"| {row.get('ID', '')} | `{left_primer}` | `{right_primer}` | {product_size} | {quality} |")
            
            lines.append("")
        
        # 可视化文件
        if visualization_files:
            lines.append("## 4. 可视化文件")
            lines.append("")
            
            for viz_name, viz_path in visualization_files.items():
                # 获取相对路径
                try:
                    rel_path = Path(viz_path).relative_to(output_dir)
                    lines.append(f"- **{viz_name}**: ![{viz_name}]({rel_path})")
                except:
                    lines.append(f"- **{viz_name}**: {viz_path}")
            
            lines.append("")
        
        lines.append("---")
        lines.append("")
        lines.append("*报告由 QTL Primer Toolkit 生成*")
        lines.append("*仅供科研参考*")
        
        # 写入文件
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))
        
        logger.info(f"Markdown报告已生成: {md_file}")
        return str(md_file)
    
    def generate_json_report(
        self,
        indels_df: Optional[pd.DataFrame],
        primers_df: Optional[pd.DataFrame],
        blast_results_df: Optional[pd.DataFrame],
        visualization_files: Optional[Dict],
        output_dir: Union[str, Path]
    ) -> Optional[str]:
        """
        生成JSON报告
        
        Args:
            indels_df: Indel数据框
            primers_df: 引物数据框
            blast_results_df: BLAST结果数据框
            visualization_files: 可视化文件路径字典
            output_dir: 输出目录
            
        Returns:
            str: JSON报告文件路径
        """
        output_dir = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_file = output_dir / f"qtl_primer_report_{timestamp}.json"
        
        # 构建报告数据
        report_data = {
            'metadata': {
                'report_type': 'QTL Primer Design Report',
                'generation_time': datetime.now().isoformat(),
                'tool_version': '1.0.0',
            },
            'statistics': {},
            'data_summaries': {},
            'visualization_files': visualization_files or {}
        }
        
        # 添加统计信息
        stats = report_data['statistics']
        
        if indels_df is not None:
            stats['indel_count'] = len(indels_df)
        
        if primers_df is not None:
            stats['primer_count'] = len(primers_df)
            if 'QUALITY' in primers_df.columns:
                stats['primer_quality_counts'] = primers_df['QUALITY'].value_counts().to_dict()
        
        if blast_results_df is not None:
            stats['blast_count'] = len(blast_results_df)
            if 'Specificity_Grade' in blast_results_df.columns:
                stats['specificity_grade_counts'] = blast_results_df['Specificity_Grade'].value_counts().to_dict()
            if 'Specificity_Score' in blast_results_df.columns:
                stats['specificity_score_stats'] = {
                    'mean': float(blast_results_df['Specificity_Score'].mean()),
                    'max': float(blast_results_df['Specificity_Score'].max()),
                    'min': float(blast_results_df['Specificity_Score'].min()),
                    'std': float(blast_results_df['Specificity_Score'].std())
                }
        
        # 添加数据摘要
        summaries = report_data['data_summaries']
        
        if indels_df is not None and not indels_df.empty:
            summaries['indels'] = {
                'row_count': len(indels_df),
                'column_count': len(indels_df.columns),
                'columns': list(indels_df.columns),
                'sample_data': indels_df.head(5).to_dict('records')
            }
        
        if primers_df is not None and not primers_df.empty:
            summaries['primers'] = {
                'row_count': len(primers_df),
                'column_count': len(primers_df.columns),
                'columns': list(primers_df.columns),
                'sample_data': primers_df.head(5).to_dict('records')
            }
        
        if blast_results_df is not None and not blast_results_df.empty:
            summaries['blast_results'] = {
                'row_count': len(blast_results_df),
                'column_count': len(blast_results_df.columns),
                'columns': list(blast_results_df.columns),
                'sample_data': blast_results_df.head(5).to_dict('records')
            }
        
        # 写入JSON文件
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"JSON报告已生成: {json_file}")
        return str(json_file)
    
    def generate_summary_report(
        self,
        indels_df: Optional[pd.DataFrame],
        primers_df: Optional[pd.DataFrame],
        blast_results_df: Optional[pd.DataFrame],
        output_dir: Union[str, Path]
    ) -> Optional[str]:
        """
        生成文本摘要报告
        
        Args:
            indels_df: Indel数据框
            primers_df: 引物数据框
            blast_results_df: BLAST结果数据框
            output_dir: 输出目录
            
        Returns:
            str: 摘要报告文件路径
        """
        output_dir = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        summary_file = output_dir / f"qtl_primer_summary_{timestamp}.txt"
        
        lines = []
        lines.append("=" * 60)
        lines.append("QTL精细定位引物设计 - 结果摘要")
        lines.append("=" * 60)
        lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        
        # 统计信息
        lines.append("统计信息:")
        lines.append("-" * 40)
        
        if indels_df is not None:
            lines.append(f"  候选Indel数量: {len(indels_df)}")
        
        if primers_df is not None:
            successful_primers = len(primers_df[primers_df['QUALITY'] != 'FAILED']) if 'QUALITY' in primers_df.columns else len(primers_df)
            lines.append(f"  设计引物数量: {len(primers_df)} (成功: {successful_primers})")
            
            if 'QUALITY' in primers_df.columns:
                quality_counts = primers_df['QUALITY'].value_counts()
                for quality, count in quality_counts.items():
                    lines.append(f"    {quality}: {count}")
        
        if blast_results_df is not None:
            lines.append(f"  BLAST验证数量: {len(blast_results_df)}")
            
            if 'Specificity_Grade' in blast_results_df.columns:
                grade_counts = blast_results_df['Specificity_Grade'].value_counts()
                lines.append("  特异性等级分布:")
                for grade, count in grade_counts.items():
                    percentage = (count / len(blast_results_df)) * 100
                    lines.append(f"    {grade}: {count} ({percentage:.1f}%)")
        
        lines.append("")
        
        # 最佳引物推荐
        if blast_results_df is not None and not blast_results_df.empty:
            lines.append("最佳引物推荐:")
            lines.append("-" * 40)
            
            # 按特异性评分排序
            if 'Specificity_Score' in blast_results_df.columns:
                top_primers = blast_results_df.nlargest(5, 'Specificity_Score')
                
                for idx, (_, row) in enumerate(top_primers.iterrows(), 1):
                    primer_name = row.get('Primer_Name', f'Primer_{idx}')
                    score = row.get('Specificity_Score', 0)
                    grade = row.get('Specificity_Grade', 'N/A')
                    
                    lines.append(f"  {idx}. {primer_name}")
                    lines.append(f"     特异性评分: {score:.1f} ({grade})")
                    
                    if 'Exact_Hits' in row:
                        lines.append(f"     完全匹配数: {row['Exact_Hits']}")
                    
                    lines.append("")
        
        lines.append("=" * 60)
        lines.append("报告结束")
        lines.append("=" * 60)
        
        # 写入文件
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))
        
        logger.info(f"摘要报告已生成: {summary_file}")
        return str(summary_file)


# 便捷函数
def generate_excel_report(
    indels_df: Optional[pd.DataFrame] = None,
    primers_df: Optional[pd.DataFrame] = None,
    blast_results_df: Optional[pd.DataFrame] = None,
    output_dir: Union[str, Path] = "./reports",
    config: Optional[Dict] = None
) -> Optional[str]:
    """
    生成Excel报告（便捷函数）
    
    Args:
        indels_df: Indel数据框
        primers_df: 引物数据框
        blast_results_df: BLAST结果数据框
        output_dir: 输出目录
        config: 配置字典
        
    Returns:
        str: Excel报告文件路径
    """
    reporter = ReportGenerator(config)
    return reporter.generate_excel_report(indels_df, primers_df, blast_results_df, output_dir)


def generate_html_report(
    indels_df: Optional[pd.DataFrame] = None,
    primers_df: Optional[pd.DataFrame] = None,
    blast_results_df: Optional[pd.DataFrame] = None,
    visualization_files: Optional[Dict] = None,
    output_dir: Union[str, Path] = "./reports",
    config: Optional[Dict] = None
) -> Optional[str]:
    """
    生成HTML报告（便捷函数）
    
    Args:
        indels_df: Indel数据框
        primers_df: 引物数据框
        blast_results_df: BLAST结果数据框
        visualization_files: 可视化文件路径字典
        output_dir: 输出目录
        config: 配置字典
        
    Returns:
        str: HTML报告文件路径
    """
    reporter = ReportGenerator(config)
    return reporter.generate_html_report(
        indels_df, primers_df, blast_results_df, 
        visualization_files, output_dir
    )


def generate_markdown_report(
    indels_df: Optional[pd.DataFrame] = None,
    primers_df: Optional[pd.DataFrame] = None,
    blast_results_df: Optional[pd.DataFrame] = None,
    visualization_files: Optional[Dict] = None,
    output_dir: Union[str, Path] = "./reports",
    config: Optional[Dict] = None
) -> Optional[str]:
    """
    生成Markdown报告（便捷函数）
    
    Args:
        indels_df: Indel数据框
        primers_df: 引物数据框
        blast_results_df: BLAST结果数据框
        visualization_files: 可视化文件路径字典
        output_dir: 输出目录
        config: 配置字典
        
    Returns:
        str: Markdown报告文件路径
    """
    reporter = ReportGenerator(config)
    return reporter.generate_markdown_report(
        indels_df, primers_df, blast_results_df,
        visualization_files, output_dir
    )
